import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
import time
from openpyxl.utils import get_column_letter


excel_sheet_gaeste = "/Users/johannesglock/Library/Mobile Documents/com~apple~CloudDocs/getraenke.xlsx"
excel_sheet_gaeste_new = "/Users/johannesglock/Library/Mobile Documents/com~apple~CloudDocs/getraenke.xlsm"

df_gaeste = pd.read_excel(excel_sheet_gaeste, sheet_name="Tabelle1", usecols="A:O", header=1, nrows=20)

st.set_page_config(page_title="Hüttentool")
st.title("Getränkeabrechnung")
st.sidebar.header("Gäste")
choice = st.sidebar.selectbox("Wähle deinen Namen:", (df_gaeste), index=1)
st.write("Du hast", choice, "ausgewählt")

wb = load_workbook(filename=excel_sheet_gaeste, read_only=False, keep_vba=True)
ws = wb.active
ws_sheet = wb["Tabelle1"]
ws_sheet.title = "Getraenke"

with st.form(key="my_form"):
    c1, c2, c3 = st.columns(3)

    with c1:
        mineral = st.number_input('Mineralwasser', step=1, key="wasser")
    with c2:
        apfel = st.number_input('Apfelschorle', step=1, key="apfel")
    with c3:
        limo = st.number_input('Limo', step=1, key="limo")

    d1, d2 = st.columns(2)

    with d1:
        holunder = st.number_input('Holunder', step=1, key="holunder")
    with d2:
        Apfelsaft = st.number_input('Apfelsaft', step=1, key="Apfelsaft")

    getraenke_liste = []
    getraenke_liste.append(mineral)
    getraenke_liste.append(apfel)
    getraenke_liste.append(limo)
    getraenke_liste.append(holunder)
    getraenke_liste.append(Apfelsaft)

    button = st.form_submit_button("Änderungen speichern")

    wb = load_workbook(filename=excel_sheet_gaeste, read_only=False, keep_vba=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=3, min_col=3, max_row=3, max_col=7):
        for cell, wert in zip(row, getraenke_liste):
            cell.value = wert
            if button:
                with st.spinner('Wait for it...'):
                    wb.save("/Users/johannesglock/Library/Mobile Documents/com~apple~CloudDocs/getraenke.xlsm")
                    st.success('Done!')

    df_gaeste_new = pd.read_excel(excel_sheet_gaeste_new, sheet_name="Tabelle1", usecols="B:O", header=None, nrows=20)
    st.dataframe(df_gaeste_new, use_container_width=True)

if button:
    st.experimental_rerun()
