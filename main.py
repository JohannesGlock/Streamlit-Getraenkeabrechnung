import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook


st.set_page_config(page_title="Hüttentool")

with open("styles.css") as f:
    st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)

excel_sheet_gaeste_new = "/Users/johannesglock/Library/Mobile Documents/com~apple~CloudDocs/getraenke.xlsm"
df_gaeste = pd.read_excel(excel_sheet_gaeste_new, sheet_name="Tabelle1", usecols="B:O", header=1, nrows=20)

st.title("Getränkeabrechnung")
st.sidebar.header("Gäste")
choice = st.sidebar.selectbox("Wähle deinen Namen:", (df_gaeste), index=1)
st.write("Du hast", choice, "ausgewählt")

wb = load_workbook(filename=excel_sheet_gaeste_new, read_only=False, keep_vba=True)
ws = wb.active
#ws_sheet = wb["Tabelle1"]
#ws_sheet.title = "Getraenke"

with st.form(key="my_form", clear_on_submit=True):
    st.write("Getränke")
    c1, c2, c3 = st.columns(3)

    with c1:
        mineral = st.number_input('Mineralwasser -still/spritzig / 0,7L', step=1, key="wasser")
    with c2:
        apfel = st.number_input('Apfelschorle / 0,5 L', step=1, key="apfel")
    with c3:
        limo = st.number_input('Limo - Orange, Zitrone, Cola-Mix / 0,5L', step=1, key="limo")

    d1, d2, d3 = st.columns(3)

    with d1:
        holunder = st.number_input('Holunder- Johannesschorle / 0,5L', step=1, key="holunder")
    with d2:
        Apfelsaft = st.number_input('Apfelsaft/Orangensaft / 1L ', step=1, key="Apfelsaft")
    with d3:
        Bier = st.number_input('Bier - alle Sorten / Flasche :beer:', step=1, key="Bier")

    e1, e2, e3 = st.columns(3)

    with e1:
        Schorle = st.number_input('Schorle Rot oder Weiß / 0,25L  :wine_glass: ', step=1, key="Schorle")
    with e2:
        Wein = st.number_input('Wein Rot, Rose oder Weiß / 0,25L', step=1, key="Wein")
    with e3:
        Wein_groß = st.number_input('Wein Rot, Rose oder Weiß / 1L', step=1, key="Wein groß")

    f1, f2, f3 = st.columns(3)

    with f1:
        Acolon_klein = st.number_input('Acolon / 0,25L', step=1, key="Acolon_groß")
    with f2:
        Acolon_groß = st.number_input('Acolon / 0,75L', step=1, key="Acolon_klein")
    with f3:
        Sekt = st.number_input('Sekt / Flasche :champagne:', step=1, key="Sekt")

    g1, g2, g3 = st.columns(3)

    with g1:
        Prosecco = st.number_input('Prosecco / Flasche', step=1, key="Prosecco")
    with g2:
        Prinz = st.number_input('Prinz(Marille, Williams...) :zany_face:', step=1, key="Prinz")
    with g3:
        Ramazotti = st.number_input('Ramazotti', step=1, key="Ramma")

    #getränke in Liste speichern
    getraenke_liste = []
    getraenke_liste.append(mineral)
    getraenke_liste.append(apfel)
    getraenke_liste.append(limo)
    getraenke_liste.append(holunder)
    getraenke_liste.append(Apfelsaft)
    getraenke_liste.append(Bier)
    getraenke_liste.append(Schorle)
    getraenke_liste.append(Wein)
    getraenke_liste.append(Wein_groß)
    getraenke_liste.append(Acolon_klein)
    getraenke_liste.append(Acolon_groß)
    getraenke_liste.append(Sekt)
    getraenke_liste.append(Prosecco)
    getraenke_liste.append(Prinz)
    getraenke_liste.append(Ramazotti)

    button = st.form_submit_button("Änderungen speichern")

    namen = []

    for row in ws.iter_rows(min_row=3, min_col=2, max_row=20, max_col=2):
        for cell in row:
            namen.append(cell.internal_value)

    #Richtige Zeile im Excel sheet wird ausgewählt und mit den angaben aus den Input felder befüllt

    if choice == namen[0]:
        for row in ws.iter_rows(min_row=3, min_col=3, max_row=3, max_col=17):
            for cell, wert in zip(row, getraenke_liste):
                cell.value += wert

    elif choice == namen[1]:
        for row in ws.iter_rows(min_row=4, min_col=3, max_row=4, max_col=17):
            for cell, wert in zip(row, getraenke_liste):
                cell.value += wert

    elif choice == namen[2]:
        for row in ws.iter_rows(min_row=5, min_col=3, max_row=5, max_col=17):
            for cell, wert in zip(row, getraenke_liste):
                cell.value += wert

    elif choice == namen[3]:
        for row in ws.iter_rows(min_row=6, min_col=3, max_row=6, max_col=17):
            for cell, wert in zip(row, getraenke_liste):
                cell.value += wert

    elif choice == namen[4]:
        for row in ws.iter_rows(min_row=7, min_col=3, max_row=7, max_col=17):
            for cell, wert in zip(row, getraenke_liste):
                cell.value += wert

    elif choice == namen[5]:
        for row in ws.iter_rows(min_row=8, min_col=3, max_row=8, max_col=17):
            for cell, wert in zip(row, getraenke_liste):
                cell.value += wert

    elif choice == namen[6]:
        for row in ws.iter_rows(min_row=9, min_col=3, max_row=9, max_col=17):
            for cell, wert in zip(row, getraenke_liste):
                cell.value += wert

    if button:
        with st.spinner('Wait for it...'):
            wb.save("/Users/johannesglock/Library/Mobile Documents/com~apple~CloudDocs/getraenke.xlsm")
            st.success('Done!')


show_button = st.button("Zeige aktuelle Liste", help="Dieser Button zeigt dir die aktuelle Liste an Getränken")
if show_button:
    df_gaeste_new = pd.read_excel(excel_sheet_gaeste_new, sheet_name="Tabelle1", usecols="B:O", header=None, nrows=20)
    st.dataframe(df_gaeste_new, use_container_width=True)
